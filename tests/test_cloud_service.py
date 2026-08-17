from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from cloud_service import (
    UploadValidationError,
    UploadedFileData,
    process_uploads,
    upload_fingerprint,
    validate_upload_metadata,
    validate_uploads,
)


def workbook_bytes():
    output = BytesIO()
    workbook = Workbook()

    height_sheet = workbook.active
    height_sheet.title = "护栏高度"
    height_sheet.append(
        [
            "地市",
            "路线编号",
            "方向",
            "原始桩号",
            "护栏类型",
            "梁板中心高度-离地(mm)",
            "路缘石高度-离地(mm)",
            "梁板中心高度(mm)",
            "异常标记",
        ]
    )
    height_sheet.append(["广州", "G15", "上行", "K1+000", "三波护栏", 0, 0, 0, ""])

    bolt_sheet = workbook.create_sheet("螺栓缺失")
    bolt_sheet.append(
        [
            "地市",
            "路线编号",
            "方向",
            "原始桩号",
            "护栏类型",
            "备注标记",
            "拼接螺栓缺失数量（颗）",
            "连接螺栓缺失数量（颗）",
        ]
    )
    bolt_sheet.append(["广州", "G15", "上行", "K1+010", "三波护栏", "", -1, 0])
    workbook.save(output)
    return output.getvalue()


def blank_workbook_bytes():
    output = BytesIO()
    Workbook().save(output)
    return output.getvalue()


def corrupted_workbook_bytes():
    output = BytesIO()
    with ZipFile(BytesIO(workbook_bytes()), "r") as source, ZipFile(
        output, "w", compression=ZIP_DEFLATED
    ) as target:
        for entry in source.infolist():
            content = source.read(entry.filename)
            if entry.filename == "xl/workbook.xml":
                content = b"<broken"
            target.writestr(entry, content)
    return output.getvalue()


def test_validate_uploads_rejects_unsupported_extension():
    uploads = [UploadedFileData("说明.txt", b"not an excel file")]

    with pytest.raises(UploadValidationError, match="仅支持"):
        validate_uploads(uploads)


def test_validate_upload_metadata_rejects_before_content_is_read():
    with pytest.raises(UploadValidationError, match="最多上传 2 个"):
        validate_upload_metadata([1, 1, 1], max_files=2)

    with pytest.raises(UploadValidationError, match="单文件不能超过 4 MiB"):
        validate_upload_metadata([5 * 1024 * 1024], max_file_bytes=4 * 1024 * 1024)

    with pytest.raises(UploadValidationError, match="总大小不能超过 8 MiB"):
        validate_upload_metadata(
            [5 * 1024 * 1024, 5 * 1024 * 1024],
            max_total_bytes=8 * 1024 * 1024,
        )


def test_validate_uploads_rejects_too_many_files():
    content = workbook_bytes()
    uploads = [UploadedFileData(f"file-{index}.xlsx", content) for index in range(3)]

    with pytest.raises(UploadValidationError, match="最多上传 2 个"):
        validate_uploads(uploads, max_files=2)


def test_validate_uploads_rejects_fake_excel_archive():
    uploads = [UploadedFileData("伪造文件.xlsx", b"not an office archive")]

    with pytest.raises(UploadValidationError, match="不是有效的 Excel"):
        validate_uploads(uploads)


def test_validate_uploads_rejects_aggregate_uncompressed_size():
    content = workbook_bytes()
    uploads = [
        UploadedFileData("first.xlsx", content),
        UploadedFileData("second.xlsx", content),
    ]

    with pytest.raises(UploadValidationError, match="累计解压体积"):
        validate_uploads(uploads, max_total_uncompressed_bytes=1)


def test_validate_uploads_sanitizes_paths_and_deduplicates_names():
    content = workbook_bytes()
    uploads = [
        UploadedFileData("../../sample.xlsx", content),
        UploadedFileData("sample.xlsx", content),
    ]

    validated = validate_uploads(uploads)

    assert [item.name for item in validated] == ["sample.xlsx", "sample_2.xlsx"]


def test_upload_fingerprint_changes_when_name_or_content_changes():
    baseline = [UploadedFileData("sample.xlsx", b"first")]

    assert upload_fingerprint(baseline) == upload_fingerprint(baseline)
    assert upload_fingerprint(baseline) != upload_fingerprint(
        [UploadedFileData("renamed.xlsx", b"first")]
    )
    assert upload_fingerprint(baseline) != upload_fingerprint(
        [UploadedFileData("sample.xlsx", b"second")]
    )


def test_process_uploads_rejects_when_every_file_fails():
    uploads = [UploadedFileData("corrupted.xlsx", corrupted_workbook_bytes())]

    with pytest.raises(UploadValidationError, match="所有文件处理失败"):
        process_uploads(uploads)


def test_process_uploads_rejects_when_no_business_sheet_is_recognized():
    uploads = [UploadedFileData("blank.xlsx", blank_workbook_bytes())]

    with pytest.raises(UploadValidationError, match="未识别到护栏高度或螺栓缺失"):
        process_uploads(uploads)


def test_process_uploads_returns_preview_excel_and_both_modified_variants():
    result = process_uploads([UploadedFileData("sample.xlsx", workbook_bytes())])

    result_workbook = load_workbook(BytesIO(result.result_excel), read_only=True)
    try:
        assert result_workbook.sheetnames == ["护栏高度", "螺栓缺失"]
    finally:
        result_workbook.close()

    assert {row["数据类型"] for row in result.preview_rows} == {"护栏高度", "螺栓缺失"}
    assert result.stats["files_success"] == 1
    assert result.result_filename.endswith(".xlsx")
    assert result.modified_filename.endswith(".zip")

    with ZipFile(BytesIO(result.modified_zip)) as archive:
        names = set(archive.namelist())
    assert "已修改数据/标黄版/sample.xlsx" in names
    assert "已修改数据/不标黄版/sample.xlsx" in names
    assert not any(name.startswith("护栏数据异常检查结果_") for name in names)
