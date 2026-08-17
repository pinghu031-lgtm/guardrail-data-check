# -*- coding: utf-8 -*-
"""护栏数据异常检查工具 v1.7 的无界面业务核心。"""

import os
import re
import zipfile
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

APP_VERSION = "v1.7"
APP_TITLE = f"护栏数据异常检查工具 {APP_VERSION}"
SUPPORTED_EXTS = (".xlsx", ".xlsm")
MODIFIED_FOLDER_NAME = "已修改数据"
MODIFIED_HIGHLIGHT_FOLDER_NAME = "标黄版"
MODIFIED_PLAIN_FOLDER_NAME = "不标黄版"

# 护栏高度异常类型。
ISSUE_MISSING_ABNORMAL_MARK_REMARK = "缺少异常标记备注"
ISSUE_WRONG_REMARK = "备注错误"
ISSUE_TRANSITION_REMARK = "过渡护栏标注错误"
ISSUE_MISSING_TRANSITION_REMARK = "缺少过渡护栏标记"
ISSUE_OTHER_GUARDRAIL_REMARK = "其他护栏标记错误"
ISSUE_HEIGHT_VALUE_ERROR = "高度值错误"
ISSUE_ZERO_VALUE_MARK_ERROR = "零值标记错误"

# 螺栓缺失异常类型。
ISSUE_BOLT_WRONG_REMARK = "螺栓备注错误"
ISSUE_BOLT_ABNORMAL_VALUE = "异常数值"

ISSUE_ORDER = {
    ISSUE_MISSING_ABNORMAL_MARK_REMARK: 1,
    ISSUE_TRANSITION_REMARK: 2,
    ISSUE_MISSING_TRANSITION_REMARK: 3,
    ISSUE_HEIGHT_VALUE_ERROR: 4,
    ISSUE_ZERO_VALUE_MARK_ERROR: 5,
    ISSUE_WRONG_REMARK: 6,
    ISSUE_OTHER_GUARDRAIL_REMARK: 7,
    ISSUE_BOLT_WRONG_REMARK: 8,
    ISSUE_BOLT_ABNORMAL_VALUE: 9,
}

CATEGORY_HEIGHT = "护栏高度"
CATEGORY_BOLT = "螺栓缺失"

BOLT_GUARDRAIL_TYPES = {"三波护栏", "两波护栏"}
BOLT_WRONG_REMARKS = {"桥梁地段", "无护栏", "其他"}
BACKGROUND_GUARDRAIL_TYPE = "背景"
TRANSITION_REMARKS = {"双层两波护栏", "双层护栏"}
TRANSITION_CONTEXT_REMARKS = TRANSITION_REMARKS | {"过渡护栏"}
TRUE_DOUBLE_LAYER_MIN_ROWS = 4
DEFAULT_ABNORMAL_MARK_REMARK = "匝道（仅路侧可用）"
INVALID_NEIGHBOR_FILL_MARKS = {"过渡护栏", "无备注"}
HEIGHT_VALUE_ERROR_IGNORED_MARKS = {"过渡护栏", "端头", "地锚端头", "双层护栏", "双层两波护栏", "植被遮挡"}
HEIGHT_VALUE_KEYS = ("height_beam_ground", "height_curb_ground", "height_beam")
ZERO_GUARDRAIL_TYPE = "-"
MODIFIED_CELL_FILL = PatternFill("solid", fgColor="FFF2CC")

COLUMN_ALIASES = {
    "city": ["地市"],
    "route": ["路线编号", "路线", "路线名称"],
    "direction": ["方向", "检测方向"],
    "pile_electronic": ["电子修正桩号"],
    "pile_marked": ["标注修正桩号"],
    "pile_original": ["原始桩号"],
    "pile_general": ["桩号", "桩号Long", "修正桩号"],
    "guardrail_type": ["护栏类型"],
    "height_beam_ground": ["梁板中心高度-离地(mm)", "梁板中心高度-离地（mm）"],
    "height_curb_ground": ["路缘石高度-离地(mm)", "路缘石高度-离地（mm）"],
    "height_beam": ["梁板中心高度(mm)", "梁板中心高度（mm）"],
    "abnormal_mark": ["异常标记", "异常备注", "异常标注", "备注"],
    "bolt_remark": ["备注标记", "螺栓备注标记", "备注"],
    "bolt_splice_missing_count": ["拼接螺栓缺失数量（颗）", "拼接螺栓缺失数量(颗)", "拼接螺栓缺失数量"],
    "bolt_connection_missing_count": ["连接螺栓缺失数量（颗）", "连接螺栓缺失数量(颗)", "连接螺栓缺失数量"],
}

HEIGHT_REQUIRED_KEYS = [
    "guardrail_type",
    "height_beam_ground",
    "height_curb_ground",
    "height_beam",
    "abnormal_mark",
]
BOLT_REQUIRED_KEYS = ["guardrail_type", "bolt_remark"]
BOLT_VALUE_KEYS = ("bolt_splice_missing_count", "bolt_connection_missing_count")
BOLT_VALUE_REQUIRED_KEYS = list(BOLT_VALUE_KEYS)

OUTPUT_HEADERS = ["序号", "地市", "路线编号", "方向", "起始桩号", "结束桩号", "异常类型"]


@dataclass
class DataRow:
    file_seq: int
    sheet_seq: int
    file_path: str
    sheet_name: str
    excel_row: int
    values: Tuple[Any, ...]
    cols: Dict[str, int]
    fallback_route: str = ""
    fallback_direction: str = ""

    def get(self, key: str) -> Any:
        idx = self.cols.get(key)
        if idx is None or idx < 0 or idx >= len(self.values):
            return None
        return self.values[idx]

    @property
    def city(self) -> str:
        return clean_display_text(self.get("city"))

    @property
    def route(self) -> str:
        return clean_display_text(self.get("route")) or self.fallback_route

    @property
    def direction(self) -> str:
        return clean_display_text(self.get("direction")) or self.fallback_direction

    @property
    def pile(self) -> str:
        # 所有异常区间定位统一使用“原始桩号”列，避免修正桩号影响起止桩号。
        return clean_display_text(self.get("pile_original"))


@dataclass
class RowIssue:
    category: str
    file_seq: int
    sheet_seq: int
    file_path: str
    sheet_name: str
    excel_row: int
    city: str
    route: str
    direction: str
    pile: str
    issue_type: str

    @property
    def group_key(self) -> Tuple[Any, ...]:
        return (
            self.category,
            self.file_seq,
            self.sheet_seq,
            self.file_path,
            self.sheet_name,
            self.city,
            self.route,
            self.direction,
            self.issue_type,
        )


@dataclass
class IssueGroup:
    category: str
    file_seq: int
    sheet_seq: int
    start_excel_row: int
    end_excel_row: int
    city: str
    route: str
    direction: str
    start_pile: str
    end_pile: str
    issue_type: str
    row_count: int

    def to_output_row(self, seq: int) -> List[Any]:
        end_pile = "-" if self.row_count <= 1 else (self.end_pile or "-")
        return [seq, self.city, self.route, self.direction, self.start_pile, end_pile, self.issue_type]


def canonical_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("\n", "").replace("\r", "")
    text = text.replace(" ", "").replace("　", "")
    text = text.replace("（", "(").replace("）", ")")
    return text


def compare_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace("　", " ")
    text = re.sub(r"\s+", "", text)
    return text


def clean_display_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "none":
        return ""
    if re.fullmatch(r"-?\d+\.0", text):
        return text[:-2]
    return text


def to_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def values_equal(left: Any, right: Any) -> bool:
    """比较写入前后的值；数值上相等的 0、0.0 和字符串零视为相同。"""
    left_number = to_number(left)
    right_number = to_number(right)
    if left_number is not None and right_number is not None:
        return abs(left_number - right_number) < 1e-9
    return compare_text(left) == compare_text(right)


def all_three_zero(values: List[Any]) -> bool:
    nums = [to_number(v) for v in values]
    return all(n is not None and abs(n) < 1e-9 for n in nums)


def any_non_zero(values: List[Any]) -> bool:
    nums = [to_number(v) for v in values]
    return any(n is not None and abs(n) >= 1e-9 for n in nums)


def is_blank_row(row_values: Tuple[Any, ...]) -> bool:
    return all(clean_display_text(v) == "" for v in row_values)


def parse_route_direction_from_filename(file_path: str) -> Tuple[str, str]:
    name = os.path.basename(file_path)
    direction = ""
    for token in ("上行", "下行", "左幅", "右幅", "内侧", "外侧"):
        if token in name:
            direction = token
            break
    route = ""
    match = re.search(r"([GSHXYC]\d{1,5})", name, flags=re.IGNORECASE)
    if match:
        route = match.group(1).upper()
    return route, direction


def find_columns(headers: Tuple[Any, ...]) -> Dict[str, int]:
    normalized_headers = [canonical_header(h) for h in headers]
    cols: Dict[str, int] = {}

    for key, aliases in COLUMN_ALIASES.items():
        alias_set = {canonical_header(a) for a in aliases}
        for i, header in enumerate(normalized_headers):
            if header in alias_set:
                cols[key] = i
                break
        if key in cols:
            continue

        if key == "pile_general":
            for i, header in enumerate(normalized_headers):
                if "桩号" in header:
                    cols[key] = i
                    break
    return cols


def find_header_row(ws, required_keys: List[str], max_scan_rows: int = 30) -> Tuple[Optional[int], Dict[str, int], Tuple[Any, ...]]:
    best_row_num: Optional[int] = None
    best_cols: Dict[str, int] = {}
    best_headers: Tuple[Any, ...] = tuple()
    best_score = -1

    for row_num, row_values in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_scan_rows), values_only=True),
        start=1,
    ):
        cols = find_columns(row_values)
        score = sum(1 for k in required_keys if k in cols) * 10
        score += sum(1 for k in ("city", "route", "direction", "pile_original", "pile_marked", "pile_electronic") if k in cols)
        if score > best_score:
            best_row_num = row_num
            best_cols = cols
            best_headers = row_values
            best_score = score

    if best_row_num is None or not all(k in best_cols for k in required_keys):
        return None, {}, tuple()
    return best_row_num, best_cols, best_headers




def is_bolt_sheet(ws, headers: Tuple[Any, ...]) -> bool:
    """识别螺栓缺失数据工作表。优先看工作表名称，其次看表头是否包含螺栓字段。"""
    title = compare_text(ws.title)
    if "螺栓" in title:
        return True
    normalized_headers = [canonical_header(h) for h in headers]
    return any("螺栓" in h for h in normalized_headers)


def find_bolt_header_row(ws) -> Tuple[Optional[int], Dict[str, int], Tuple[Any, ...]]:
    header_row, cols, headers = find_header_row(ws, BOLT_REQUIRED_KEYS)
    if header_row is not None:
        return header_row, cols, headers
    header_row, cols, headers = find_header_row(ws, BOLT_VALUE_REQUIRED_KEYS)
    if header_row is not None:
        return header_row, cols, headers

    for row_num, row_values in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True),
        start=1,
    ):
        normalized_headers = [canonical_header(h) for h in row_values]
        if any("螺栓" in header for header in normalized_headers):
            return row_num, find_columns(row_values), row_values
    return None, {}, tuple()

def guardrail_type_changed(data_rows: List[DataRow], idx: int) -> bool:
    """判断双层两波护栏/双层护栏备注所在行的前后护栏类型是否发生变化。"""
    current = compare_text(data_rows[idx].get("guardrail_type"))
    prev_type = compare_text(data_rows[idx - 1].get("guardrail_type")) if idx > 0 else ""
    next_type = compare_text(data_rows[idx + 1].get("guardrail_type")) if idx + 1 < len(data_rows) else ""

    if current:
        if prev_type and prev_type != current:
            return True
        if next_type and next_type != current:
            return True
        return False
    if prev_type and next_type and prev_type != next_type:
        return True
    return False


def adjacent_transition_remark(data_rows: List[DataRow], idx: int) -> bool:
    """判断当前行前一条或后一条有效数据的异常标记是否为“过渡护栏”。"""
    prev_mark = compare_text(data_rows[idx - 1].get("abnormal_mark")) if idx > 0 else ""
    next_mark = compare_text(data_rows[idx + 1].get("abnormal_mark")) if idx + 1 < len(data_rows) else ""
    return prev_mark == "过渡护栏" or next_mark == "过渡护栏"


def is_double_layer_remark_row(row: DataRow) -> bool:
    """判断当前行是否标记为双层两波护栏或双层护栏。"""
    return compare_text(row.get("abnormal_mark")) in TRANSITION_REMARKS


def are_adjacent_excel_rows(data_rows: List[DataRow], left_idx: int, right_idx: int) -> bool:
    return data_rows[right_idx].excel_row == data_rows[left_idx].excel_row + 1


def same_data_context(left: DataRow, right: DataRow, include_guardrail_type: bool = False) -> bool:
    """判断两行是否属于同一连续业务数据段。"""
    same_identity = (left.city, left.route, left.direction) == (right.city, right.route, right.direction)
    if not same_identity:
        return False
    if include_guardrail_type:
        return compare_text(left.get("guardrail_type")) == compare_text(right.get("guardrail_type"))
    return True


def are_same_data_segment(
    data_rows: List[DataRow], left_idx: int, right_idx: int, include_guardrail_type: bool = False
) -> bool:
    return are_adjacent_excel_rows(data_rows, left_idx, right_idx) and same_data_context(
        data_rows[left_idx], data_rows[right_idx], include_guardrail_type
    )


def find_true_double_layer_indices(data_rows: List[DataRow]) -> set:
    """一次扫描找出连续4行及以上的真实双层护栏行，避免逐行反复扫描。"""
    true_indices = set()
    start = 0
    while start < len(data_rows):
        if not is_double_layer_remark_row(data_rows[start]):
            start += 1
            continue
        end = start
        while (
            end + 1 < len(data_rows)
            and are_same_data_segment(data_rows, end, end + 1, include_guardrail_type=True)
            and is_double_layer_remark_row(data_rows[end + 1])
        ):
            end += 1
        if end - start + 1 >= TRUE_DOUBLE_LAYER_MIN_ROWS:
            true_indices.update(range(start, end + 1))
        start = end + 1
    return true_indices


def is_true_double_layer_remark_segment(data_rows: List[DataRow], idx: int, true_indices: Optional[set] = None) -> bool:
    """连续4行及以上双层标记视为真实双层护栏，不参与过渡标注错误判断。"""
    if true_indices is not None:
        return idx in true_indices
    if not is_double_layer_remark_row(data_rows[idx]):
        return False

    start = idx
    while (
        start > 0
        and are_same_data_segment(data_rows, start - 1, start, include_guardrail_type=True)
        and is_double_layer_remark_row(data_rows[start - 1])
    ):
        start -= 1

    end = idx
    while (
        end + 1 < len(data_rows)
        and are_same_data_segment(data_rows, end, end + 1, include_guardrail_type=True)
        and is_double_layer_remark_row(data_rows[end + 1])
    ):
        end += 1

    return end - start + 1 >= TRUE_DOUBLE_LAYER_MIN_ROWS


def is_transition_remark_error(data_rows: List[DataRow], idx: int, true_indices: Optional[set] = None) -> bool:
    """综合判断“双层护栏/双层两波护栏”是否属于过渡护栏标注错误。"""
    mark = compare_text(data_rows[idx].get("abnormal_mark"))
    if mark not in TRANSITION_REMARKS:
        return False
    if is_true_double_layer_remark_segment(data_rows, idx, true_indices):
        return False
    return guardrail_type_changed(data_rows, idx) or adjacent_transition_remark(data_rows, idx)


def is_missing_transition_remark(data_rows: List[DataRow], idx: int) -> bool:
    """
    判断当前行是否缺少“过渡护栏”标记。

    规则：
    1. 当前行必须同时存在前一条和后一条有效数据；
    2. 前一行异常标记为“双层护栏”“双层两波护栏”或“过渡护栏”；
    3. 后一行护栏类型不是“三波护栏”或“两波护栏”；
    4. 当前行异常标记必须完全为空。

    满足以上条件时，将当前行判定为“缺少过渡护栏标记”。
    """
    if idx <= 0 or idx + 1 >= len(data_rows):
        return False

    previous_mark = compare_text(data_rows[idx - 1].get("abnormal_mark"))
    current_mark = compare_text(data_rows[idx].get("abnormal_mark"))
    next_guardrail_type = compare_text(data_rows[idx + 1].get("guardrail_type"))

    if previous_mark not in TRANSITION_CONTEXT_REMARKS:
        return False
    if next_guardrail_type in BOLT_GUARDRAIL_TYPES:
        return False
    if current_mark != "":
        return False
    return True


def row_height_values(row: DataRow) -> List[Any]:
    return [row.get(key) for key in HEIGHT_VALUE_KEYS]


def all_height_values_present(values: List[Any]) -> bool:
    return all(to_number(value) is not None for value in values)


def has_effective_abnormal_mark(row: DataRow) -> bool:
    mark = compare_text(row.get("abnormal_mark"))
    return mark != "" and mark != "无备注"


def has_no_effective_abnormal_mark(row: DataRow) -> bool:
    mark = compare_text(row.get("abnormal_mark"))
    return mark == "" or mark == "无备注"


def is_height_value_error(row: DataRow) -> bool:
    """三列高度都有数值且存在非零值，同时异常标记不在忽略名单内时，判定为高度值错误。"""
    heights = row_height_values(row)
    return (
        all_height_values_present(heights)
        and any_non_zero(heights)
        and has_effective_abnormal_mark(row)
        and compare_text(row.get("abnormal_mark")) not in HEIGHT_VALUE_ERROR_IGNORED_MARKS
    )


def is_zero_value_mark_error(row: DataRow) -> bool:
    """护栏类型不为“-”，三列高度均为0，且无有效异常标记时，判定为零值标记错误。"""
    guardrail_type = compare_text(row.get("guardrail_type"))
    return (
        guardrail_type != ZERO_GUARDRAIL_TYPE
        and all_three_zero(row_height_values(row))
        and has_no_effective_abnormal_mark(row)
    )


def is_missing_abnormal_mark_remark(row: DataRow) -> bool:
    """护栏类型不是三波/两波护栏，且异常标记为空时，判定为缺少异常标记备注。"""
    guardrail_type = compare_text(row.get("guardrail_type"))
    mark = compare_text(row.get("abnormal_mark"))
    return (
        guardrail_type not in BOLT_GUARDRAIL_TYPES
        and guardrail_type != ZERO_GUARDRAIL_TYPE
        and mark == ""
        and not is_zero_value_mark_error(row)
    )


def make_issue(row: DataRow, category: str, issue_type: str) -> RowIssue:
    return RowIssue(
        category=category,
        file_seq=row.file_seq,
        sheet_seq=row.sheet_seq,
        file_path=row.file_path,
        sheet_name=row.sheet_name,
        excel_row=row.excel_row,
        city=row.city,
        route=row.route,
        direction=row.direction,
        pile=row.pile,
        issue_type=issue_type,
    )


def check_height_row_issues(data_rows: List[DataRow], idx: int, true_double_layer_indices: Optional[set] = None) -> List[RowIssue]:
    row = data_rows[idx]
    if is_missing_abnormal_mark_remark(row):
        return [make_issue(row, CATEGORY_HEIGHT, ISSUE_MISSING_ABNORMAL_MARK_REMARK)]
    if is_transition_remark_error(data_rows, idx, true_double_layer_indices):
        return [make_issue(row, CATEGORY_HEIGHT, ISSUE_TRANSITION_REMARK)]
    if is_missing_transition_remark(data_rows, idx):
        return [make_issue(row, CATEGORY_HEIGHT, ISSUE_MISSING_TRANSITION_REMARK)]
    if is_height_value_error(row):
        return [make_issue(row, CATEGORY_HEIGHT, ISSUE_HEIGHT_VALUE_ERROR)]
    if is_zero_value_mark_error(row):
        return [make_issue(row, CATEGORY_HEIGHT, ISSUE_ZERO_VALUE_MARK_ERROR)]
    return []


def check_bolt_row_issues(row: DataRow) -> List[RowIssue]:
    guardrail_type = compare_text(row.get("guardrail_type"))
    remark = compare_text(row.get("bolt_remark"))
    issues: List[RowIssue] = []
    if guardrail_type in BOLT_GUARDRAIL_TYPES and remark in BOLT_WRONG_REMARKS:
        issues.append(make_issue(row, CATEGORY_BOLT, ISSUE_BOLT_WRONG_REMARK))
    if has_negative_bolt_value(row):
        issues.append(make_issue(row, CATEGORY_BOLT, ISSUE_BOLT_ABNORMAL_VALUE))
    return issues


def has_negative_bolt_value(row: DataRow) -> bool:
    return any(is_negative_value(row.get(key)) for key in BOLT_VALUE_KEYS)


def is_negative_value(value: Any) -> bool:
    number = to_number(value)
    return number is not None and number < 0


def group_row_issues(issues: List[RowIssue]) -> List[IssueGroup]:
    if not issues:
        return []

    issues_sorted = sorted(
        issues,
        key=lambda x: (
            x.category,
            x.file_seq,
            x.sheet_seq,
            ISSUE_ORDER.get(x.issue_type, 99),
            x.excel_row,
            x.city,
            x.route,
            x.direction,
        ),
    )

    groups: List[IssueGroup] = []
    current_key: Optional[Tuple[Any, ...]] = None
    start_issue: Optional[RowIssue] = None
    last_issue: Optional[RowIssue] = None
    count = 0

    def flush_group() -> None:
        nonlocal start_issue, last_issue, count
        if start_issue is None or last_issue is None:
            return
        groups.append(
            IssueGroup(
                category=start_issue.category,
                file_seq=start_issue.file_seq,
                sheet_seq=start_issue.sheet_seq,
                start_excel_row=start_issue.excel_row,
                end_excel_row=last_issue.excel_row,
                city=start_issue.city,
                route=start_issue.route,
                direction=start_issue.direction,
                start_pile=start_issue.pile,
                end_pile=last_issue.pile,
                issue_type=start_issue.issue_type,
                row_count=count,
            )
        )

    for issue in issues_sorted:
        key = issue.group_key
        is_continuous = current_key == key and last_issue is not None and issue.excel_row == last_issue.excel_row + 1
        if not is_continuous:
            flush_group()
            current_key = key
            start_issue = issue
            last_issue = issue
            count = 1
        else:
            last_issue = issue
            count += 1

    flush_group()
    groups.sort(key=lambda g: (g.file_seq, g.sheet_seq, g.start_excel_row, ISSUE_ORDER.get(g.issue_type, 99)))
    return groups


def build_data_rows(file_seq: int, sheet_seq: int, file_path: str, ws, header_row: int, cols: Dict[str, int], fallback_route: str, fallback_direction: str) -> List[DataRow]:
    data_rows: List[DataRow] = []
    for excel_row, row_values in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if is_blank_row(row_values):
            continue
        data_rows.append(
            DataRow(
                file_seq=file_seq,
                sheet_seq=sheet_seq,
                file_path=file_path,
                sheet_name=ws.title,
                excel_row=excel_row,
                values=row_values,
                cols=cols,
                fallback_route=fallback_route,
                fallback_direction=fallback_direction,
            )
        )
    return data_rows


def load_workbook_for_inspection(file_path: str):
    """以兼容模式读取检查数据；现场文件的工作表维度元数据可能不准确。"""
    return load_workbook(file_path, read_only=False, data_only=True)


def load_workbook_for_modification(file_path: str):
    """打开待修改工作簿；处理 xlsm 时保留 VBA 工程。"""
    keep_vba = file_path.lower().endswith(".xlsm")
    return load_workbook(file_path, keep_vba=keep_vba)


def load_workbook_for_cached_values(file_path: str):
    """兼容读取公式缓存值，不依赖可能错误的工作表维度元数据。"""
    return load_workbook(file_path, read_only=False, data_only=True)


def workbook_has_formulas(file_path: str) -> bool:
    """快速扫描 xlsx 内工作表 XML，判断是否存在公式单元格。"""
    try:
        with zipfile.ZipFile(file_path) as zf:
            for name in zf.namelist():
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    data = zf.read(name)
                    if b"<f>" in data or b"<f " in data:
                        return True
    except Exception:
        return True
    return False





class InspectionProcessor:
    def __init__(self, logger=None):
        self.logger = logger

    def log(self, message: str) -> None:
        if self.logger:
            self.logger(message)

    def process_files(self, file_paths: List[str]) -> Tuple[List[IssueGroup], Dict[str, Any]]:
        all_issues: List[RowIssue] = []
        stats = {
            "files_total": len(file_paths),
            "files_success": 0,
            "files_failed": 0,
            "height_sheets_processed": 0,
            "bolt_sheets_processed": 0,
            "sheets_skipped": 0,
            "errors": [],
        }

        for file_seq, file_path in enumerate(file_paths, start=1):
            try:
                self.log(f"正在读取：{file_path}")
                issues = self.process_one_file(file_seq, file_path, stats)
                all_issues.extend(issues)
                stats["files_success"] += 1
                self.log(f"完成：{os.path.basename(file_path)}，发现明细异常 {len(issues)} 条。")
            except Exception as exc:
                stats["files_failed"] += 1
                error_text = f"{file_path}：{exc}"
                stats["errors"].append(error_text)
                self.log(f"处理失败：{error_text}")
                traceback.print_exc()

        groups = group_row_issues(all_issues)
        return groups, stats

    def process_one_file(self, file_seq: int, file_path: str, stats: Dict[str, Any]) -> List[RowIssue]:
        wb = load_workbook_for_inspection(file_path)
        try:
            fallback_route, fallback_direction = parse_route_direction_from_filename(file_path)
            file_issues: List[RowIssue] = []

            for sheet_seq, ws in enumerate(wb.worksheets, start=1):
                handled = False

                height_header_row, height_cols, _ = find_header_row(ws, HEIGHT_REQUIRED_KEYS)
                if height_header_row is not None:
                    handled = True
                    stats["height_sheets_processed"] += 1
                    data_rows = build_data_rows(file_seq, sheet_seq, file_path, ws, height_header_row, height_cols, fallback_route, fallback_direction)
                    true_double_layer_indices = find_true_double_layer_indices(data_rows)
                    self.log(f"  处理护栏高度工作表：{ws.title}，数据行 {len(data_rows)} 行。")
                    for idx in range(len(data_rows)):
                        file_issues.extend(check_height_row_issues(data_rows, idx, true_double_layer_indices))

                bolt_header_row, bolt_cols, bolt_headers = find_bolt_header_row(ws)
                # 含高度字段的工作表里也可能有“备注”列，避免误识别为螺栓表。
                if bolt_header_row is not None and is_bolt_sheet(ws, bolt_headers):
                    handled = True
                    stats["bolt_sheets_processed"] += 1
                    data_rows = build_data_rows(file_seq, sheet_seq, file_path, ws, bolt_header_row, bolt_cols, fallback_route, fallback_direction)
                    self.log(f"  处理螺栓缺失工作表：{ws.title}，数据行 {len(data_rows)} 行。")
                    for row in data_rows:
                        file_issues.extend(check_bolt_row_issues(row))

                if not handled:
                    stats["sheets_skipped"] += 1
                    self.log(f"  跳过工作表：{ws.title}（未识别到护栏高度或螺栓缺失所需字段）")

            return file_issues
        finally:
            wb.close()


def style_output_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(OUTPUT_HEADERS)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = {"A": 8, "B": 14, "C": 14, "D": 10, "E": 18, "F": 18, "G": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_one_category_sheet(wb: Workbook, title: str, groups: List[IssueGroup]) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(OUTPUT_HEADERS)
    for seq, group in enumerate(groups, start=1):
        ws.append(group.to_output_row(seq))
    if not groups:
        ws.append(["", "", "", "", "", "", "未发现符合规则的异常"])
    style_output_sheet(ws)


def export_groups_to_excel(groups: List[IssueGroup], output_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    height_groups = [g for g in groups if g.category == CATEGORY_HEIGHT]
    bolt_groups = [g for g in groups if g.category == CATEGORY_BOLT]
    export_one_category_sheet(wb, "护栏高度", height_groups)
    export_one_category_sheet(wb, "螺栓缺失", bolt_groups)
    wb.save(output_path)


def cell_value(ws, excel_row: int, cols: Dict[str, int], key: str) -> Any:
    idx = cols.get(key)
    if idx is None:
        return None
    return ws.cell(row=excel_row, column=idx + 1).value


def modified_cell_fill(highlight_modified: bool):
    return MODIFIED_CELL_FILL if highlight_modified else None


def set_cell_value(ws, excel_row: int, cols: Dict[str, int], key: str, value: Any, fill=None, record: Optional[list] = None) -> None:
    idx = cols.get(key)
    if idx is None:
        return
    cell = ws.cell(row=excel_row, column=idx + 1)
    cell.value = value
    if fill is not None:
        cell.fill = fill
    if record is not None:
        record.append((ws.title, excel_row, idx + 1))


def set_height_values_to_zero(ws, excel_row: int, cols: Dict[str, int], fill=None, record: Optional[list] = None) -> int:
    modified = 0
    for height_key in HEIGHT_VALUE_KEYS:
        old_value = cell_value(ws, excel_row, cols, height_key)
        if values_equal(old_value, 0):
            continue
        set_cell_value(ws, excel_row, cols, height_key, 0, fill=fill, record=record)
        modified += 1
    return modified


def is_valid_neighbor_fill_mark(mark: Any) -> bool:
    text = clean_display_text(mark)
    return bool(text) and compare_text(text) not in INVALID_NEIGHBOR_FILL_MARKS


def missing_abnormal_fill_values(data_rows: List[DataRow]) -> Dict[int, str]:
    """仅使用同一连续业务段中、缺失区间紧邻两侧的有效备注。"""
    result: Dict[int, str] = {}
    segment_start = 0
    while segment_start < len(data_rows):
        segment_end = segment_start
        while segment_end + 1 < len(data_rows) and are_same_data_segment(
            data_rows, segment_end, segment_end + 1
        ):
            segment_end += 1

        idx = segment_start
        while idx <= segment_end:
            if not is_missing_abnormal_mark_remark(data_rows[idx]):
                idx += 1
                continue

            missing_start = idx
            while idx + 1 <= segment_end and is_missing_abnormal_mark_remark(data_rows[idx + 1]):
                idx += 1
            missing_end = idx

            previous_item: Optional[Tuple[str, int]] = None
            if missing_start > segment_start:
                mark = clean_display_text(data_rows[missing_start - 1].get("abnormal_mark"))
                if is_valid_neighbor_fill_mark(mark):
                    previous_item = (mark, missing_start - 1)

            next_item: Optional[Tuple[str, int]] = None
            if missing_end < segment_end:
                mark = clean_display_text(data_rows[missing_end + 1].get("abnormal_mark"))
                if is_valid_neighbor_fill_mark(mark):
                    next_item = (mark, missing_end + 1)

            for missing_idx in range(missing_start, missing_end + 1):
                if previous_item and next_item:
                    previous_distance = missing_idx - previous_item[1]
                    next_distance = next_item[1] - missing_idx
                    fill_value = (
                        previous_item[0]
                        if previous_distance <= next_distance
                        else next_item[0]
                    )
                elif previous_item:
                    fill_value = previous_item[0]
                elif next_item:
                    fill_value = next_item[0]
                else:
                    fill_value = DEFAULT_ABNORMAL_MARK_REMARK
                result[data_rows[missing_idx].excel_row] = fill_value

            idx += 1
        segment_start = segment_end + 1
    return result


def fill_missing_abnormal_mark_remarks(
    ws,
    data_rows: List[DataRow],
    cols: Dict[str, int],
    skip_excel_rows: Optional[set] = None,
    highlight_modified: bool = False,
    record: Optional[list] = None,
) -> int:
    """为缺少异常标记备注的行补充备注，并同步将三列高度值置为0。"""
    modified = 0
    skip_excel_rows = skip_excel_rows or set()
    fill = modified_cell_fill(highlight_modified)
    fill_values = missing_abnormal_fill_values(data_rows)
    for row in data_rows:
        if row.excel_row in skip_excel_rows or row.excel_row not in fill_values:
            continue
        modified += set_cell_if_changed(
            ws, row.excel_row, cols, "abnormal_mark", fill_values[row.excel_row], fill=fill, record=record
        )
        modified += set_height_values_to_zero(ws, row.excel_row, cols, fill=fill, record=record)
    return modified


def apply_height_modifications(
    ws, header_row: int, cols: Dict[str, int], file_path: str, sheet_seq: int = 1,
    highlight_modified: bool = False, value_ws=None, record: Optional[list] = None,
) -> int:
    fallback_route, fallback_direction = parse_route_direction_from_filename(file_path)
    source_ws = value_ws or ws
    data_rows = build_data_rows(1, sheet_seq, file_path, source_ws, header_row, cols, fallback_route, fallback_direction)
    true_double_layer_indices = find_true_double_layer_indices(data_rows)
    modified = 0
    fill = modified_cell_fill(highlight_modified)

    # 先找出需要补充“过渡护栏”的行。若这些行也符合缺少异常标记备注，
    # 不再先填默认匝道备注，避免同一单元格被重复修改。
    missing_transition_rows = {
        row.excel_row
        for idx, row in enumerate(data_rows)
        if is_missing_transition_remark(data_rows, idx)
    }

    # 规则1对应修改：缺少异常标记备注，按前后行有效异常标记补充。
    modified += fill_missing_abnormal_mark_remarks(
        ws,
        data_rows,
        cols,
        skip_excel_rows=missing_transition_rows,
        highlight_modified=highlight_modified,
        record=record,
    )

    # 其他规则按原始数据判断，直接改对应异常标记。
    # 过渡类修正优先级高于清空错误备注，避免符合过渡条件的行被先清空。
    for idx, row in enumerate(data_rows):
        if is_transition_remark_error(data_rows, idx, true_double_layer_indices):
            modified += set_cell_if_changed(ws, row.excel_row, cols, "abnormal_mark", "过渡护栏", fill=fill, record=record)
            continue

        if is_missing_transition_remark(data_rows, idx):
            modified += set_cell_if_changed(ws, row.excel_row, cols, "abnormal_mark", "过渡护栏", fill=fill, record=record)
            continue

        if is_height_value_error(row):
            modified += set_height_values_to_zero(ws, row.excel_row, cols, fill=fill, record=record)
            continue

        if is_zero_value_mark_error(row):
            modified += set_cell_if_changed(ws, row.excel_row, cols, "guardrail_type", ZERO_GUARDRAIL_TYPE, fill=fill, record=record)
            continue

    return modified


def is_bolt_wrong_remark_row(row: DataRow) -> bool:
    """判断一行是否属于“螺栓备注错误”。"""
    guardrail_type = compare_text(row.get("guardrail_type"))
    remark = compare_text(row.get("bolt_remark"))
    return guardrail_type in BOLT_GUARDRAIL_TYPES and remark in BOLT_WRONG_REMARKS


def split_continuous_bolt_issue_segments(data_rows: List[DataRow]) -> List[List[int]]:
    """
    将螺栓备注错误按 Excel 连续行拆分为区间。

    返回值中的每个元素均为 data_rows 的索引列表。空白行会中断连续区间，
    从而避免把不相邻的数据误合并为同一问题区间。
    """
    issue_indices = [idx for idx, row in enumerate(data_rows) if is_bolt_wrong_remark_row(row)]
    if not issue_indices:
        return []

    segments: List[List[int]] = []
    current = [issue_indices[0]]
    for idx in issue_indices[1:]:
        previous_idx = current[-1]
        rows_are_continuous = (
            idx == previous_idx + 1
            and data_rows[idx].excel_row == data_rows[previous_idx].excel_row + 1
        )
        if rows_are_continuous:
            current.append(idx)
        else:
            segments.append(current)
            current = [idx]
    segments.append(current)
    return segments


def set_cell_if_changed(ws, excel_row: int, cols: Dict[str, int], key: str, value: Any, fill=None, record: Optional[list] = None) -> int:
    """仅在目标值确实发生变化时写入，并返回修改单元格数量。"""
    old_value = cell_value(ws, excel_row, cols, key)
    if values_equal(old_value, value):
        return 0
    idx = cols.get(key)
    if idx is None:
        return 0
    cell = ws.cell(row=excel_row, column=idx + 1)
    cell.value = value
    if fill is not None:
        cell.fill = fill
    if record is not None:
        record.append((ws.title, excel_row, idx + 1))
    return 1


def set_bolt_cell_if_changed(ws, excel_row: int, cols: Dict[str, int], key: str, value: Any, highlight_modified: bool = False, record: Optional[list] = None) -> int:
    return set_cell_if_changed(ws, excel_row, cols, key, value, fill=modified_cell_fill(highlight_modified), record=record)


def set_negative_bolt_values_to_zero(ws, row: DataRow, highlight_modified: bool = False, record: Optional[list] = None) -> int:
    modified = 0
    fill = modified_cell_fill(highlight_modified)
    for key in BOLT_VALUE_KEYS:
        if is_negative_value(row.get(key)):
            modified += set_cell_if_changed(ws, row.excel_row, row.cols, key, 0, fill=fill, record=record)
    return modified


def apply_bolt_modifications(
    ws, header_row: int, cols: Dict[str, int], file_path: str, sheet_seq: int = 1,
    highlight_modified: bool = False, value_ws=None, record: Optional[list] = None,
) -> int:
    """
    按异常区间前后护栏类型修改“螺栓备注错误”。

    1. 单行异常：
       - 前后护栏类型均为“背景”时，将当前行护栏类型改为“背景”；
       - 前后护栏类型均属于“三波护栏/两波护栏”时，清空当前行备注标记；
       - 其他无法明确判断的单行情况，沿用原逻辑清空备注标记。
    2. 连续2～3行异常：
       - 问题区间内护栏类型一致；
       - 区间前后护栏类型彼此不同；
       - 区间前后护栏类型均与问题区间护栏类型不同；
       满足以上条件时，将问题区间全部护栏类型改为“背景”。
       其他情况沿用原逻辑清空问题区间备注标记。
    3. 连续4行及以上异常：沿用原逻辑清空备注标记。
    """
    fallback_route, fallback_direction = parse_route_direction_from_filename(file_path)
    source_ws = value_ws or ws
    data_rows = build_data_rows(1, sheet_seq, file_path, source_ws, header_row, cols, fallback_route, fallback_direction)
    modified = 0

    for row in data_rows:
        modified += set_negative_bolt_values_to_zero(ws, row, highlight_modified=highlight_modified, record=record)

    for segment in split_continuous_bolt_issue_segments(data_rows):
        start_idx = segment[0]
        end_idx = segment[-1]
        segment_length = len(segment)
        previous_type = compare_text(data_rows[start_idx - 1].get("guardrail_type")) if start_idx > 0 else ""
        next_type = compare_text(data_rows[end_idx + 1].get("guardrail_type")) if end_idx + 1 < len(data_rows) else ""

        if segment_length == 1:
            row = data_rows[start_idx]
            if previous_type == BACKGROUND_GUARDRAIL_TYPE and next_type == BACKGROUND_GUARDRAIL_TYPE:
                modified += set_bolt_cell_if_changed(
                    ws, row.excel_row, cols, "guardrail_type", BACKGROUND_GUARDRAIL_TYPE, highlight_modified=highlight_modified, record=record
                )
            elif previous_type in BOLT_GUARDRAIL_TYPES and next_type in BOLT_GUARDRAIL_TYPES:
                modified += set_bolt_cell_if_changed(ws, row.excel_row, cols, "bolt_remark", "", highlight_modified=highlight_modified, record=record)
            else:
                # 边界行或前后类型不能形成明确证据时，保留v1.5的默认修正方式。
                modified += set_bolt_cell_if_changed(ws, row.excel_row, cols, "bolt_remark", "", highlight_modified=highlight_modified, record=record)
            continue

        if segment_length in (2, 3):
            segment_types = {compare_text(data_rows[idx].get("guardrail_type")) for idx in segment}
            segment_types.discard("")
            segment_type = next(iter(segment_types)) if len(segment_types) == 1 else ""

            should_change_to_background = (
                bool(segment_type)
                and bool(previous_type)
                and bool(next_type)
                and previous_type != next_type
                and previous_type != segment_type
                and next_type != segment_type
            )

            if should_change_to_background:
                for idx in segment:
                    row = data_rows[idx]
                    modified += set_bolt_cell_if_changed(
                        ws, row.excel_row, cols, "guardrail_type", BACKGROUND_GUARDRAIL_TYPE, highlight_modified=highlight_modified, record=record
                    )
            else:
                for idx in segment:
                    row = data_rows[idx]
                    modified += set_bolt_cell_if_changed(ws, row.excel_row, cols, "bolt_remark", "", highlight_modified=highlight_modified, record=record)
            continue

        # 未明确要求的连续4行及以上问题区间，继续采用原有清空备注逻辑。
        for idx in segment:
            row = data_rows[idx]
            modified += set_bolt_cell_if_changed(ws, row.excel_row, cols, "bolt_remark", "", highlight_modified=highlight_modified, record=record)

    return modified


def unique_output_path(folder: str, filename: str) -> str:
    base, ext = os.path.splitext(filename)
    candidate = os.path.join(folder, filename)
    idx = 1
    while os.path.exists(candidate):
        candidate = os.path.join(folder, f"{base}_{idx}{ext}")
        idx += 1
    return candidate


def export_modified_workbooks(file_paths: List[str], output_parent: str, logger=None) -> Tuple[str, Dict[str, Any]]:
    def log(message: str) -> None:
        if logger:
            logger(message)

    output_folder = os.path.join(output_parent, MODIFIED_FOLDER_NAME)
    highlight_output_folder = os.path.join(output_folder, MODIFIED_HIGHLIGHT_FOLDER_NAME)
    plain_output_folder = os.path.join(output_folder, MODIFIED_PLAIN_FOLDER_NAME)
    os.makedirs(highlight_output_folder, exist_ok=True)
    os.makedirs(plain_output_folder, exist_ok=True)
    stats = {"files_total": len(file_paths), "files_success": 0, "files_failed": 0, "modified_cells": 0, "errors": []}

    def export_one_file_both_variants(file_path: str) -> Tuple[int, str, str]:
        # 每个文件只加载一次源工作簿；仅当存在公式时才额外加载缓存值工作簿。
        # 处理一遍后分别保存标黄版与不标黄版。
        wb = load_workbook_for_modification(file_path)
        value_wb = None
        modified_cells: List[Tuple[str, int, int]] = []
        try:
            if workbook_has_formulas(file_path):
                value_wb = load_workbook_for_cached_values(file_path)
            file_modified = 0
            for sheet_seq, ws in enumerate(wb.worksheets, start=1):
                value_ws = value_wb.worksheets[sheet_seq - 1] if value_wb is not None else ws
                height_header_row, height_cols, _ = find_header_row(ws, HEIGHT_REQUIRED_KEYS)
                if height_header_row is not None:
                    count = apply_height_modifications(
                        ws,
                        height_header_row,
                        height_cols,
                        file_path,
                        sheet_seq,
                        highlight_modified=False,
                        value_ws=value_ws,
                        record=modified_cells,
                    )
                    file_modified += count
                    log(f"  护栏高度工作表 {ws.title} 修改 {count} 个单元格。")

                bolt_header_row, bolt_cols, bolt_headers = find_bolt_header_row(ws)
                if bolt_header_row is not None and is_bolt_sheet(ws, bolt_headers):
                    count = apply_bolt_modifications(
                        ws,
                        bolt_header_row,
                        bolt_cols,
                        file_path,
                        sheet_seq,
                        highlight_modified=False,
                        value_ws=value_ws,
                        record=modified_cells,
                    )
                    file_modified += count
                    log(f"  螺栓缺失工作表 {ws.title} 修改 {count} 个单元格。")

            # 先保存不标黄版。
            plain_path = unique_output_path(plain_output_folder, os.path.basename(file_path))
            wb.save(plain_path)

            # 给修改过的单元格设置填充色，再保存标黄版。
            for sheet_name, row, col in modified_cells:
                wb[sheet_name].cell(row=row, column=col).fill = MODIFIED_CELL_FILL
            highlight_path = unique_output_path(highlight_output_folder, os.path.basename(file_path))
            wb.save(highlight_path)
            return file_modified, plain_path, highlight_path
        finally:
            if value_wb is not None:
                value_wb.close()
            wb.close()

    max_workers = max(1, min(4, os.cpu_count() or 1))

    def process_one_file(file_path: str) -> Tuple[str, Any]:
        try:
            log(f"正在修改并导出：{file_path}")
            file_modified, plain_path, highlight_path = export_one_file_both_variants(file_path)
            log(f"完成导出：标黄版 {highlight_path}；不标黄版 {plain_path}；修改 {file_modified} 个单元格。")
            return "ok", file_modified
        except Exception as exc:
            error_text = f"{file_path}：{exc}"
            log(f"修改导出失败：{error_text}")
            traceback.print_exc()
            return "fail", error_text

    if len(file_paths) == 1:
        results = [process_one_file(file_paths[0])]
    else:
        results = []
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = [pool.submit(process_one_file, fp) for fp in file_paths]
            for future in as_completed(futures):
                results.append(future.result())

    for status, payload in results:
        if status == "ok":
            stats["files_success"] += 1
            stats["modified_cells"] += payload
        else:
            stats["files_failed"] += 1
            stats["errors"].append(payload)
    return output_folder, stats
